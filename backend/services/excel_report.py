"""
Excel report generator for bulk candidate export.

Called from routes/hr_reports.py via build_bulk_xlsx(). Returns the XLSX as
bytes (caller wraps in a StreamingResponse).

DESIGN DECISIONS

1. One row per candidate. No multi-sheet, no per-question detail. The
   point is "HR's working spreadsheet" — they sort, filter, pivot in
   Excel. A 1-row-per-candidate flat layout is the only one that supports
   all those operations.

2. ALL invitations from this HR get exported, regardless of completion
   status. HR sees in_progress / submitted / scored alongside each other.
   They can filter to "scored only" inside Excel themselves.

3. NO embedded transcripts or essays. Those don't fit Excel cells well
   (long, multiline, often with quotes). HR uses the per-candidate PDF
   for that detail.

4. Header row is frozen so HR can scroll through 200 candidates without
   losing column context. Column widths set so common values fit without
   manual resizing.

5. Rating column is conditionally formatted (green/amber/red) so HR can
   eyeball the spreadsheet for "who passed."

6. Numeric columns are formatted as numbers, not strings — so HR can
   sort them numerically (sorting strings makes 9 > 75).
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Optional
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Invitation, Score


# ------------------------------------------------------------------
# Styles
# ------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="0B2545")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Rating-specific cell fills (applied to the Rating column)
RATING_FILLS = {
    "recommended": PatternFill("solid", fgColor="D7F0DD"),       # light green
    "borderline": PatternFill("solid", fgColor="FCEFD3"),        # light amber
    "not_recommended": PatternFill("solid", fgColor="F8DADA"),   # light red
}
RATING_FONTS = {
    "recommended": Font(color="1F7A3A", bold=True),
    "borderline": Font(color="A86A00", bold=True),
    "not_recommended": Font(color="A02828", bold=True),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7E2"),
    right=Side(style="thin", color="D0D7E2"),
    top=Side(style="thin", color="D0D7E2"),
    bottom=Side(style="thin", color="D0D7E2"),
)


# ------------------------------------------------------------------
# Column definitions
# ------------------------------------------------------------------
# (header_label, column_width_chars)
# Order here = column order in the output. Width is approximate
# (Excel char-widths are not pixels; tuned by eyeballing).
COLUMNS = [
    ("Invitation ID", 12),
    ("Candidate Name", 22),
    ("Email", 28),
    ("Level", 12),
    ("Sections", 16),
    ("Status", 14),
    ("Created At", 18),
    ("Started At", 18),
    ("Submitted At", 18),
    ("Reading Score", 12),
    ("Writing Score", 12),
    ("Speaking Score", 13),
    ("Total Score", 11),
    ("Rating", 16),
    ("Tab Switches", 12),
    ("Off-Tab Seconds", 14),
    ("Submission Reason", 22),
    ("Terminated", 11),
    ("Total Time (sec)", 14),
]


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------
def _fmt_dt(
    dt: Optional[datetime],
    target_tz_name: Optional[str] = None,
    tz_label: Optional[str] = None,
) -> Optional[str]:
    """
    Render a stored UTC datetime in the candidate's local timezone.

    Same logic as services/pdf_report.py._fmt_dt — kept in sync so
    the PDF report and Excel export show identical timestamps for the
    same candidate. Each invitation has its own display_timezone, so
    this function is called per-row with that row's target zone.

    Args:
        dt: stored datetime (treated as UTC if naive)
        target_tz_name: IANA zone name like "Asia/Kolkata", or None/UTC
            for UTC display
        tz_label: short suffix like "IST" / "ET" appended in parentheses

    Failure modes (all safe — never raises):
      - dt is None              → returns None (Excel renders empty cell)
      - target_tz_name is bad   → falls back to UTC display with "(UTC)"
      - tzdata not installed    → same fallback (Windows w/o tzdata)

    Returns strings like:
      "2026-05-07 16:35 (IST)"   ← normal case
      "2026-05-07 11:05 (UTC)"   ← fallback
      None                        ← when dt is None (cell stays empty)
    """
    if dt is None:
        return None

    # Treat naive datetimes as UTC. Python 3.12+ deprecates the implicit
    # assumption, so be explicit.
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    # Default render = UTC if no target zone given
    if not target_tz_name or target_tz_name == "UTC":
        return dt.strftime("%Y-%m-%d %H:%M") + " (UTC)"

    # Convert to candidate's zone. On failure, fall back to UTC so the
    # time AND the label match — better than rendering a non-UTC time
    # with a UTC label, which would silently mislead.
    try:
        target = ZoneInfo(target_tz_name)
        local = dt.astimezone(target)
        label = tz_label or target_tz_name
        return local.strftime("%Y-%m-%d %H:%M") + f" ({label})"
    except ZoneInfoNotFoundError:
        return dt.strftime("%Y-%m-%d %H:%M") + " (UTC)"


def _status_label(inv: Invitation, score: Optional[Score]) -> str:
    """
    Compute human-readable status from the multiple state fields on Invitation.

    Priority order:
      - terminated (derived from submission_reason) → "Terminated"
      - score row exists → "Scored"
      - submitted_at set → "Submitted (scoring...)"
      - started_at set   → "In progress"
      - else             → "Not started"

    "Terminated" is derived from submission_reason rather than a stored
    boolean column. See _is_terminated for the taxonomy.
    """
    if _is_terminated(inv):
        return "Terminated"
    if score is not None:
        return "Scored"
    if inv.submitted_at:
        return "Submitted (scoring...)"
    if inv.started_at:
        return "In progress"
    return "Not started"


def _is_terminated(inv: Invitation) -> bool:
    """
    Check whether the test was forcibly ended (vs. completed normally).

    Termination is implied by submission_reason values that indicate auto-
    submission rather than the candidate clicking Finish:
      - tab_switch_termination
      - reading_timer_expired / writing_timer_expired / speaking_timer_expired
    Anything else (candidate_finished, None for in-progress) is "normal".
    """
    return bool(inv.submission_reason and inv.submission_reason not in (
        "candidate_finished", "candidate_submit",
    ))


def _sections_label(inv: Invitation) -> str:
    """Compact representation of which sections were included. e.g. 'R+W+S'."""
    parts = []
    if inv.include_reading:
        parts.append("R")
    if inv.include_writing:
        parts.append("W")
    if inv.include_speaking:
        parts.append("S")
    return "+".join(parts) if parts else "—"


def _rating_label(rating: Optional[str]) -> str:
    """Human label for the rating column. Empty string for unscored rows."""
    if rating == "recommended":
        return "Recommended"
    if rating == "borderline":
        return "Borderline"
    if rating == "not_recommended":
        return "Not Recommended"
    return ""


def _total_test_seconds(inv: Invitation) -> Optional[int]:
    """Calculate total test time in seconds from started_at to submitted_at."""
    if inv.started_at and inv.submitted_at:
        return int((inv.submitted_at - inv.started_at).total_seconds())
    return None


# ------------------------------------------------------------------
# Main entry point
# ------------------------------------------------------------------
def build_bulk_xlsx(
    invitations_with_scores: list[tuple[Invitation, Optional[Score]]],
    tz_label_map: Optional[dict[str, str]] = None,
    hr_label_map: Optional[dict[int, str]] = None,
) -> bytes:
    """
    Generate a bulk-export XLSX file containing candidates.

    Args:
        invitations_with_scores: list of (Invitation, Score|None) tuples.
            The caller is responsible for the join — keeps DB access in the
            route, not the formatter.
        tz_label_map: optional dict mapping IANA zone name → short label
            (e.g. {"Asia/Kolkata": "IST", "America/New_York": "ET"}).
            Looked up from the supported_timezones table by the caller.
            When a candidate's display_timezone has no entry here, the IANA
            name is used as the label (e.g. "Asia/Kolkata" itself).
            When None, all timestamps render in UTC.
        hr_label_map: optional dict mapping inv.hr_admin_id → human-readable
            label (e.g. {7: "Sinchana Gowda <sinchana@stixis.com>"}).
            When provided, an "HR Admin" column is inserted as the second
            column (between Invitation ID and Candidate Name) and populated
            from this map. When None, no HR column is added — the export
            stays single-HR-scoped, identical to the existing HR dashboard
            export.

            This is what makes the same builder useful for both:
              - HR's own export (single HR scope, no HR column needed)
              - Admin's all-candidates export (multi-HR scope, HR column
                tells the admin who sent each invitation)

    Per-row timezone is taken from inv.display_timezone — same field the
    PDF report uses, same zone HR picked when sending the invitation.
    Different candidates can have different timezones; this function
    handles each one independently.

    Returns the XLSX as bytes. Caller wraps in StreamingResponse.
    """
    tz_label_map = tz_label_map or {}
    include_hr_column = hr_label_map is not None
    hr_label_map = hr_label_map or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    # Build the column list dynamically. The base COLUMNS is shared across
    # both single-HR and multi-HR exports. When the admin needs the
    # multi-HR version, we splice in an "HR Admin" column right after
    # "Invitation ID" so the HR identity sits at the start of each row,
    # next to the invitation key. Putting it any later would force the
    # eye to scan past several columns of candidate data before learning
    # whose invitation it is.
    columns = list(COLUMNS)
    if include_hr_column:
        columns.insert(1, ("HR Admin", 30))

    # ---- Header row ----
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    # Freeze the header so it stays visible while scrolling
    ws.freeze_panes = "A2"

    # ---- Data rows ----
    for row_offset, (inv, score) in enumerate(invitations_with_scores, start=2):
        # Resolve THIS invitation's display timezone. Different candidates
        # can have different zones — HR picks per-invitation. We translate
        # IANA name → short label here so each datetime cell can show
        # "(IST)" / "(ET)" / "(UTC)" alongside the time.
        candidate_tz_name = inv.display_timezone or "UTC"
        candidate_tz_label = (
            tz_label_map.get(candidate_tz_name) or candidate_tz_name
            if candidate_tz_name and candidate_tz_name != "UTC"
            else None
        )

        # Build the row values in the same order as `columns` above.
        # The HR Admin value is inserted via list splicing below so the
        # base values list stays identical to the single-HR case — easier
        # to keep in sync with COLUMNS over time.
        values = [
            inv.id,
            inv.candidate_name,
            inv.candidate_email,
            (inv.difficulty or "").title(),
            _sections_label(inv),
            _status_label(inv, score),
            _fmt_dt(inv.created_at, candidate_tz_name, candidate_tz_label),
            _fmt_dt(inv.started_at, candidate_tz_name, candidate_tz_label),
            _fmt_dt(inv.submitted_at, candidate_tz_name, candidate_tz_label),
            score.reading_score if score else None,
            score.writing_score if score else None,
            score.speaking_score if score else None,
            score.total_score if score else None,
            _rating_label(score.rating) if score else "",
            inv.tab_switches_count or 0,
            inv.tab_switches_total_seconds or 0,
            inv.submission_reason or "",
            "Yes" if _is_terminated(inv) else "No",
            _total_test_seconds(inv),
        ]

        # Splice the HR Admin value into position 1 (right after Invitation ID)
        # only when the HR column is enabled. Done here rather than in the
        # values literal above so the base 19-element list stays unchanged
        # — minimises drift risk if more columns are added later.
        if include_hr_column:
            values.insert(1, hr_label_map.get(inv.hr_admin_id, "—"))

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=val)
            cell.border = THIN_BORDER

            # Column-specific styling. Use `columns` (the local list that
            # may include "HR Admin"), not COLUMNS (the module constant)
            # — otherwise the styling indexes shift by one when HR column
            # is present and Rating/Sections/etc. style the wrong cells.
            label = columns[col_idx - 1][0]
            if label in ("Reading Score", "Writing Score", "Speaking Score",
                         "Total Score", "Tab Switches", "Off-Tab Seconds",
                         "Total Time (sec)", "Invitation ID"):
                # Numeric columns: right-aligned, integer format
                cell.alignment = Alignment(horizontal="right")
                if isinstance(val, int):
                    cell.number_format = "0"
            elif label == "Rating":
                # Conditional fill based on rating value
                if score and score.rating in RATING_FILLS:
                    cell.fill = RATING_FILLS[score.rating]
                    cell.font = RATING_FONTS[score.rating]
                cell.alignment = Alignment(horizontal="center")
            elif label in ("Sections", "Terminated", "Status"):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # ---- Auto-filter on the header row ----
    # This adds the dropdown arrows on each column header in Excel,
    # letting HR/admin filter without manually selecting the range.
    # Uses `columns` (local) so the range covers the HR column too when
    # present — using COLUMNS would leave the HR column un-filterable.
    last_col = get_column_letter(len(columns))
    last_row = len(invitations_with_scores) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # ---- Output as bytes ----
    buf = BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    buf.close()
    return xlsx_bytes